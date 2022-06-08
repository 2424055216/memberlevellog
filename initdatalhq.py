import pymysql
import openpyxl

wb = openpyxl.load_workbook('testmemberlevellog.xlsx')
ws = wb.active

# 此方法不可用，null数据执行sql处理会有问题

db = pymysql.connect(host="localhost", user="root", passwd="12345678", db="db")

cursor = db.cursor()
# sql = """insert into em_h5_record (id,name,url,member_id,key1,key2,open_id,page_id,url_param,content_json,
# del_flag,creator,creator_id,created_at,updator,updator_id,updated_at,privacy_policy)
# values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) """
# id,member_id,tenant_id,level_template_id,level_id,\
# level_name,from_level,to_level,description,extra_json,\
# maintain,occurred_at,created_at,updated_at,dl_auto_created_at,\
# dl_auto_updated_at,cdc_hash_code

# id,member_id,tenant_id,level_template_id,level_id,level_name,from_level,to_level,description,extra_json,maintain,occurred_at,created_at,updated_at

for i in range(0, 128):
    sql1 = "insert into mc_level_change_log_" + str(i)
    sql2 = """(id,member_id,tenant_id,level_template_id,level_id,level_name,from_level,to_level,description,extra_json,maintain,occurred_at,created_at,updated_at
) values """

    valueses = []
    m = 0
    for j in range(1, ws.max_row+1):
        j = str(j)
        # print("表"+ws['A' + j].value)
        # print("INSERT INTO `mc_level_change_log_" + str(i) + "` VALUES(null")
        if ws['A' + j].value == "INSERT INTO `mc_level_change_log_" + str(i) + "` VALUES(null":
            m += 1  # 需要处理null数据
            value = (m, ws['B' + j].value, int(), ws['D' + j].value, ws['E' + j].value, ws['F' + j].value,
                     ws['G' + j].value, ws['H' + j].value, ws['I' + j].value, ws['J' + j].value, ws['K' + j].value, ws['L' + j].value, ws['M' + j].value,
                     ws['N' + j].value)
            valueses.append(value)
    s = str(valueses)
    s = s.lstrip('[')
    s = s.rstrip(']')
    if len(s) > 0:
        sql = sql1 + sql2 + s
        print(sql)
        cursor.execute(sql)

cursor.close()
db.commit()
db.close()
